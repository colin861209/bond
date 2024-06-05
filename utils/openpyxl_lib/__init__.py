from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.styles import Alignment
from datetime import datetime
import os

# Sheet number
SHEET_DATA_NAME =          '淨值'
SHEET_REFERWEBSITE_NAME =  '參考債券網站'
SHEET_SUMMARY_NAME =       '投資總覽'

file_name = f'{os.getcwd()}\\投資明細_Colin.xlsm'

excel_fix_name = ['日期', '美元/台幣', '買進匯率', '賣出匯率']
def returnExcelColumnLetter(num:int) -> str: return get_column_letter(num)
def returnExcelColumnIndex(col:str) -> int: return column_index_from_string(col)
def return_strA(num=excel_fix_name.index('日期')+1) -> str: return returnExcelColumnLetter(num)
def return_strB(num=excel_fix_name.index('美元/台幣')+1) -> str: return returnExcelColumnLetter(num)
def return_strC(num=excel_fix_name.index('買進匯率')+1) -> str: return returnExcelColumnLetter(num)
def return_strD(num=excel_fix_name.index('賣出匯率')+1) -> str: return returnExcelColumnLetter(num)
str_A=return_strA()
str_B=return_strB()
str_C=return_strC()
str_D=return_strD()


def returnVal(target): return target.value

def writeVal(address, target, align:str='') -> bool: 
    if target is float: address.number_format = numbers.FORMAT_NUMBER_00
    if align == "center": address.alignment=Alignment(horizontal='center', vertical='center')
    if target is datetime: address.number_format = numbers.FORMAT_DATE_YYYYMMDD2
    address.value = target

def open_workbook(sheet_name, file_name=file_name):
    wb=load_workbook(file_name, read_only=False, keep_vba=True)
    ws=wb[sheet_name]
    return wb, ws